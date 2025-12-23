import pandas as pd # type: ignore
import psycopg  # type: ignore
import re
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from app.db.connection import connection_to_db # Assuming this is your connection module


def generate_semester_attendance_report_xls(sem_id: str, output_path: str):
    """
    Generates a comprehensive consolidated Excel attendance report for a semester.

    Creates a single-sheet report with:
    - Rows: Students
    - Columns: Student ID, Student Name, then each course showing "Present/Total (Percentage)"
    - Color-coded percentages for quick identification of low attendance

    This format is more readable for university administration compared to
    multiple sheets per course.

    Args:
        sem_id (str): The ID of the semester to generate the report for.
        output_path (str): The path to save the output .xlsx file.
    """
    print(f"🚀 Starting consolidated attendance report for semester: {sem_id}")

    # Query to get semester and program info
    info_sql = """
        SELECT
            sem.sem_name,
            p.prog_name
        FROM semester sem
        JOIN program p ON sem.prog_id = p.prog_id
        WHERE sem.sem_id = %(sem_id)s
        LIMIT 1;
    """

    # Query to get all students enrolled in courses for this semester
    # This includes both semester students (via student_enrollment) and local/backlog students (via course_students)
    sql = """
        WITH semester_courses AS (
            -- Get all courses in this semester
            SELECT course_id, course_name
            FROM course
            WHERE sem_id = %(sem_id)s
        ),
        all_course_students AS (
            -- Get students enrolled in semester (they have access to all courses)
            SELECT DISTINCT
                s.student_id,
                s.student_name,
                sc.course_id,
                sc.course_name,
                TRUE as is_enrolled
            FROM students s
            JOIN student_enrollment se ON s.student_id = se.student_id
            CROSS JOIN semester_courses sc
            WHERE se.sem_id = %(sem_id)s

            UNION

            -- Get local/backlog students (only enrolled in specific courses via course_students)
            SELECT DISTINCT
                s.student_id,
                s.student_name,
                cs.course_id,
                sc.course_name,
                TRUE as is_enrolled
            FROM students s
            JOIN course_students cs ON s.student_id = cs.student_id
            JOIN semester_courses sc ON cs.course_id = sc.course_id
            WHERE NOT EXISTS (
                -- Exclude students already enrolled in semester
                SELECT 1 FROM student_enrollment se
                WHERE se.student_id = s.student_id AND se.sem_id = %(sem_id)s
            )
        ),
        all_students_all_courses AS (
            -- Create a cross join of all unique students with all courses
            -- to ensure we have N/A entries for courses a student isn't enrolled in
            SELECT DISTINCT
                acs.student_id,
                acs.student_name,
                sc.course_id,
                sc.course_name,
                CASE
                    WHEN EXISTS (
                        SELECT 1 FROM all_course_students acs2
                        WHERE acs2.student_id = acs.student_id
                        AND acs2.course_id = sc.course_id
                    ) THEN TRUE
                    ELSE FALSE
                END as is_enrolled
            FROM (SELECT DISTINCT student_id, student_name FROM all_course_students) acs
            CROSS JOIN semester_courses sc
        )
        SELECT
            asac.student_id,
            asac.student_name,
            asac.course_id,
            asac.course_name,
            asac.is_enrolled,
            COALESCE(COUNT(*) FILTER (WHERE a.present = true), 0) AS present_count,
            COALESCE(COUNT(a.present), 0) AS total_classes
        FROM all_students_all_courses asac
        LEFT JOIN attendance a ON a.student_id = asac.student_id AND a.course_id = asac.course_id
        GROUP BY asac.student_id, asac.student_name, asac.course_id, asac.course_name, asac.is_enrolled
        ORDER BY asac.student_name, asac.course_name;
    """

    try:
        with connection_to_db() as conn:
            print("✅ Database connection successful.")

            # Get semester and program info
            info_df = pd.read_sql_query(info_sql, conn, params={'sem_id': sem_id})
            sem_name = info_df['sem_name'].iloc[0] if not info_df.empty else sem_id
            prog_name = info_df['prog_name'].iloc[0] if not info_df.empty else "Unknown Program"

            df = pd.read_sql_query(sql, conn, params={'sem_id': sem_id})
            print(f"Found {len(df)} student-course attendance records.")

        if df.empty:
            print(f"⚠️ No attendance data found for semester '{sem_id}'.")
            # Create empty report with message
            empty_df = pd.DataFrame({'Message': ['No attendance data found for this semester']})
            empty_df.to_excel(output_path, index=False)
            return

        # Calculate percentage for each row
        df['percentage'] = df.apply(
            lambda row: round((row['present_count'] / row['total_classes'] * 100), 1)
            if row['total_classes'] > 0 else 0, axis=1
        )

        # Create formatted attendance string: "Present/Total (XX%)" or "N/A" if not enrolled
        df['attendance_str'] = df.apply(
            lambda row: "N/A" if not row['is_enrolled']
            else (f"{row['present_count']}/{row['total_classes']} ({row['percentage']}%)"
                  if row['total_classes'] > 0 else "0/0 (0%)"),
            axis=1
        )

        # Create column header with Course Name (Course ID)
        df['course_header'] = df.apply(
            lambda row: f"{row['course_name']} ({row['course_id']})",
            axis=1
        )

        # Pivot to get courses as columns
        pivot_df = df.pivot_table(
            index=['student_id', 'student_name'],
            columns='course_header',
            values='attendance_str',
            aggfunc='first',
            fill_value='N/A'
        )

        pivot_df.reset_index(inplace=True)

        # Rename columns for clarity
        pivot_df.rename(columns={
            'student_id': 'Student ID',
            'student_name': 'Student Name'
        }, inplace=True)

        # Also create a percentage pivot for color coding
        percent_pivot = df.pivot_table(
            index=['student_id', 'student_name'],
            columns='course_header',
            values='percentage',
            aggfunc='first',
            fill_value=0
        )
        percent_pivot.reset_index(inplace=True)

        # Reorder columns: Student ID, Student Name, then courses (alphabetically sorted)
        course_cols = [col for col in pivot_df.columns if col not in ['Student ID', 'Student Name']]
        ordered_cols = ['Student ID', 'Student Name'] + sorted(course_cols)
        pivot_df = pivot_df[ordered_cols]

        # Calculate overall attendance for each student (only for courses they're enrolled in)
        def calculate_overall(row):
            total_present = 0
            total_classes = 0
            for col in course_cols:
                cell_val = row[col]
                # Skip N/A cells (student not enrolled in this course)
                if cell_val and cell_val != 'N/A' and '/' in str(cell_val):
                    try:
                        # Parse "Present/Total (XX%)" format
                        parts = str(cell_val).split('/')
                        present = int(parts[0])
                        total = int(parts[1].split(' ')[0])
                        total_present += present
                        total_classes += total
                    except (ValueError, IndexError):
                        pass
            if total_classes > 0:
                percentage = round((total_present / total_classes * 100), 1)
                return f"{total_present}/{total_classes} ({percentage}%)"
            return "N/A"

        pivot_df['Overall Attendance'] = pivot_df.apply(calculate_overall, axis=1)

        # Update ordered columns to include Overall Attendance at the end
        ordered_cols = ['Student ID', 'Student Name'] + sorted(course_cols) + ['Overall Attendance']
        pivot_df = pivot_df[ordered_cols]

        # Write to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write main data
            pivot_df.to_excel(
                writer, sheet_name="Semester Attendance", index=False, startrow=3
            )

            worksheet = writer.sheets["Semester Attendance"]

            # Add title
            worksheet.merge_cells('A1:F1')
            title_cell = worksheet['A1']
            title_cell.value = "SEMESTER ATTENDANCE REPORT"
            title_cell.font = Font(bold=True, size=18)
            title_cell.alignment = Alignment(horizontal='center')

            # Add subtitle with semester and program info
            worksheet.merge_cells('A2:F2')
            subtitle_cell = worksheet['A2']
            subtitle_cell.value = f"Semester: {sem_name} | Program: {prog_name} | Format: Present/Total (Percentage)"
            subtitle_cell.font = Font(bold=True, size=11)
            subtitle_cell.alignment = Alignment(horizontal='center')

            # Style the header row
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for cell in worksheet[4]:  # Row 4 is the header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = thin_border

            # Define color fills
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            orange_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # For N/A cells

            # Style data rows and apply conditional formatting
            for row_idx in range(5, worksheet.max_row + 1):
                for col_idx, cell in enumerate(worksheet[row_idx], 1):
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')

                    # Color code cells
                    if cell.value and isinstance(cell.value, str):
                        cell_str = str(cell.value)
                        # Gray for N/A cells (student not enrolled in course)
                        if cell_str == 'N/A':
                            cell.fill = gray_fill
                        # Color code cells containing percentages
                        elif '%' in cell_str:
                            try:
                                # Extract percentage from string like "10/15 (66.7%)"
                                percent_str = cell_str.split('(')[1].replace('%)', '')
                                percent_val = float(percent_str)
                                if percent_val >= 90:
                                    cell.fill = green_fill
                                elif percent_val >= 75:
                                    cell.fill = yellow_fill
                                elif percent_val >= 60:
                                    cell.fill = orange_fill
                                else:
                                    cell.fill = red_fill
                            except (ValueError, IndexError):
                                pass

            # Auto-adjust column widths (skip merged cells)
            for col_idx in range(1, worksheet.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for row_idx in range(4, worksheet.max_row + 1):  # Start from header row (4)
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if not isinstance(cell, MergedCell) and cell.value:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 20)
                worksheet.column_dimensions[col_letter].width = adjusted_width

            # Set specific column widths
            worksheet.column_dimensions['A'].width = 15  # Student ID
            worksheet.column_dimensions['B'].width = 25  # Student Name

        print(f"\n🎉 Consolidated report saved to '{os.path.abspath(output_path)}'")

    except psycopg.Error as e:
        print(f"❌ Database Error: {e}")
        raise
    except Exception as e:
        print(f"❌ An unexpected error occurred: {e}")
        raise