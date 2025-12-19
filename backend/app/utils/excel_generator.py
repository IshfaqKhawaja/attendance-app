import pandas as pd
from typing import Optional, Dict
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


def generate_attendance_excel(df: pd.DataFrame, output_path: str = "attendance_report.xlsx", course_info: Optional[Dict[str, str]] = None) -> Optional[str]:
    """
    Generates a comprehensive pivoted attendance report supporting multiple daily sessions.

    Args:
        df (pd.DataFrame): Attendance data with columns:
                           student_id, student_name, date, lecture_num, status ('P' or 'A').
        output_path (str): File path to save the Excel report.
        course_info (dict, optional): Course information with keys:
                           course_id, course_name, sem_name, prog_name.

    Note: Only counts classes where the student has a record (P or A).
          N/A values are excluded from total count.
    """
    if df.empty:
        print("⚠️ No data available to generate report.")
        return None

    # Pivot the data: each date+lecture becomes a column
    pivot_df = df.pivot_table(
        index=["student_id", "student_name"],
        columns=['date', 'lecture_num'],
        values='status',
        aggfunc='first',
        fill_value='N/A'  # N/A for days when student wasn't enrolled
    )

    # Format column headers as "DD/MM/YYYY (Lec N)"
    if isinstance(pivot_df.columns, pd.MultiIndex):
        pivot_df.columns = [
            f"{date.strftime('%d/%m/%Y')} (L{lec_num})"
            for date, lec_num in pivot_df.columns
        ]

    pivot_df.reset_index(inplace=True)

    # Get attendance columns
    attendance_cols = [col for col in pivot_df.columns if col not in ['student_id', 'student_name']]

    # Calculate Present count (only count 'P', ignore 'N/A')
    pivot_df['Present'] = pivot_df[attendance_cols].apply(
        lambda row: (row == 'P').sum(), axis=1
    )

    # Calculate Total classes PER STUDENT (exclude N/A - only count where student has records)
    # This fixes the issue where students who joined later don't get penalized for days they weren't enrolled
    pivot_df['Total'] = pivot_df[attendance_cols].apply(
        lambda row: ((row == 'P') | (row == 'A')).sum(), axis=1
    )

    # Calculate Absent count
    pivot_df['Absent'] = pivot_df['Total'] - pivot_df['Present']

    # Calculate percentage based on student's actual total
    pivot_df['Percentage'] = pivot_df.apply(
        lambda row: round((row['Present'] / row['Total'] * 100), 1) if row['Total'] > 0 else 0,
        axis=1
    )

    # Reorder columns: student info first, then summary, then daily attendance
    summary_cols = ['student_id', 'student_name', 'Present', 'Absent', 'Total', 'Percentage']
    ordered_cols = summary_cols + attendance_cols
    pivot_df = pivot_df[ordered_cols]

    # Rename columns for better readability
    pivot_df.rename(columns={
        'student_id': 'Student ID',
        'student_name': 'Student Name'
    }, inplace=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Start data at row 6 to leave room for title and course info
        pivot_df.to_excel(writer, sheet_name="Attendance Report", index=False, startrow=5)

        worksheet = writer.sheets["Attendance Report"]

        # Add title
        worksheet.merge_cells('A1:F1')
        title_cell = worksheet['A1']
        title_cell.value = "ATTENDANCE REPORT"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        # Add course info on separate rows for better readability
        if course_info:
            # Row 2: Course Name and ID
            worksheet.merge_cells('A2:F2')
            course_cell = worksheet['A2']
            course_cell.value = f"Course: {course_info.get('course_name', 'N/A')} ({course_info.get('course_id', 'N/A')})"
            course_cell.font = Font(size=11, bold=True)
            course_cell.alignment = Alignment(horizontal='center')

            # Row 3: Semester
            worksheet.merge_cells('A3:F3')
            sem_cell = worksheet['A3']
            sem_cell.value = f"Semester: {course_info.get('sem_name', 'N/A')}"
            sem_cell.font = Font(size=10)
            sem_cell.alignment = Alignment(horizontal='center')

            # Row 4: Program
            worksheet.merge_cells('A4:F4')
            prog_cell = worksheet['A4']
            prog_cell.value = f"Program: {course_info.get('prog_name', 'N/A')}"
            prog_cell.font = Font(size=10)
            prog_cell.alignment = Alignment(horizontal='center')
        else:
            worksheet.merge_cells('A2:F2')
            subtitle_cell = worksheet['A2']
            subtitle_cell.value = "Course Attendance Details"
            subtitle_cell.font = Font(size=10, italic=True)
            subtitle_cell.alignment = Alignment(horizontal='center')

        # Style the header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in worksheet[6]:  # Row 6 is the header (after title and course info)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Style data rows and apply conditional formatting for percentage
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        orange_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Find the percentage column index
        percentage_col = None
        for idx, cell in enumerate(worksheet[6], 1):
            if cell.value == 'Percentage':
                percentage_col = idx
                break

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=7, max_row=worksheet.max_row), 7):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            # Color code the percentage cell
            if percentage_col:
                percent_cell = row[percentage_col - 1]
                try:
                    percent_val = float(percent_cell.value) if percent_cell.value else 0
                    if percent_val >= 90:
                        percent_cell.fill = green_fill
                    elif percent_val >= 75:
                        percent_cell.fill = yellow_fill
                    elif percent_val >= 60:
                        percent_cell.fill = orange_fill
                    else:
                        percent_cell.fill = red_fill
                except (ValueError, TypeError):
                    pass

        # Auto-adjust column widths (skip merged cells)
        for col_idx in range(1, worksheet.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(6, worksheet.max_row + 1):  # Start from header row (6)
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell) and cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 15)  # Cap width at 15
            worksheet.column_dimensions[col_letter].width = adjusted_width

        # Set student name column wider
        worksheet.column_dimensions['B'].width = 25

    print(f"✅ Excel report successfully generated at '{output_path}'")
    return output_path